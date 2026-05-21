from __future__ import annotations

import io

import pytest

from app.services import csv_parser


def _make_orders_detailed_workbook():
    try:
        import openpyxl
    except Exception:  # pragma: no cover - optional dependency
        pytest.skip("openpyxl not installed")

    wb = openpyxl.Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Generated report"])

    orders = wb.create_sheet("Orders Detailed")
    for _ in range(6):
        orders.append([None])

    headers = [
        "Order Ref",
        "Buyer",
        "Buyer Account ID",
        "For Delivery",
        "Delivery Address",
        "Note",
        "Product",
        "Product Code",
        "Unit Type",
        "Price",
        "Quantity Ordered",
        "Net",
    ]
    orders.append(headers)

    rows = [
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 1",
            "Oaty Breakfast Cookie",
            "OB-1",
            "unit",
            13.90,
            1,
            13.90,
        ],
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 2",
            "Ricciarelli (Lemon cookies without gluten)",
            "C179",
            "unit",
            18.00,
            2,
            36.00,
        ],
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 3",
            "Single Layer Vegan Chocolate Cake 10''",
            "CAKE-10",
            "unit",
            32.00,
            1,
            32.00,
        ],
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 4",
            "Biscoff Blondie",
            "BL-1",
            "unit",
            25.60,
            1,
            25.60,
        ],
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 5",
            "Chocolate Biscuit Cake - LOTTS & CO SHALLOW VERSION",
            "CBC-1",
            "unit",
            38.00,
            1,
            38.00,
        ],
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 6",
            "Selection Box of 6",
            "SB-6",
            "unit",
            10.50,
            2,
            21.00,
        ],
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 7",
            "Biscoff Blondie Box of 4 slices",
            "BB-4",
            "unit",
            6.00,
            2,
            12.00,
        ],
        [
            "4481118",
            "Example Buyer",
            "B-448",
            "2026-04-02",
            "1 Main St",
            "Note 8",
            "Selection Box of 4",
            "SB-4",
            "unit",
            7.30,
            2,
            14.60,
        ],
        [
            "4453118",
            "Lotts & Co - Clontarf",
            "B-445",
            "2026-04-02",
            "2 Main St",
            "Note 9",
            "Lotts & Co sample order",
            "L-445",
            "unit",
            367.60,
            1,
            367.60,
        ],
    ]

    for row in rows:
        orders.append(row)

    buyers = wb.create_sheet("Buyers")
    buyers.append(["Buyer", "Order Quantity", "Net", "Total"])
    buyers.append(["Lotts & Co - Clontarf", 1, 367.60, 417.23])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_parse_orders_detailed_uses_product_name_sku_quantity_and_net():
    content = _make_orders_detailed_workbook()
    orders, debug = csv_parser.parse_unify_file_bytes(content, "unify-report.xlsx")

    assert debug["source"] == "csv"
    assert len(orders) == 2

    order_448 = next(order for order in orders if order["order_id"] == "4481118")
    assert order_448["customer_name"] == "Example Buyer"
    assert order_448["buyer_id"] == "B-448"
    assert order_448["preview_status"] == "ready"
    assert order_448["preview_reason"] == ""
    assert order_448["total"] == pytest.approx(193.10)
    assert len(order_448["lines"]) == 8

    ricciarelli = next(line for line in order_448["lines"] if line["item_sku"] == "C179")
    assert ricciarelli["item_name"] == "Ricciarelli (Lemon cookies without gluten)"
    assert ricciarelli["item_sku"] == "C179"
    assert ricciarelli["quantity"] == pytest.approx(2.0)
    assert ricciarelli["price"] == pytest.approx(18.0)
    assert ricciarelli["quantity"] * ricciarelli["price"] == pytest.approx(36.0)

    order_445 = next(order for order in orders if order["order_id"] == "4453118")
    assert order_445["customer_name"] == "Lotts & Co - Clontarf"
    assert order_445["total"] == pytest.approx(367.60)
    assert order_445["preview_status"] == "ready"


def test_parse_orders_detailed_header_row_can_start_on_row_7():
    content = _make_orders_detailed_workbook()
    orders, _debug = csv_parser.parse_unify_file_bytes(content, "unify-report.xlsx")

    assert orders[0]["order_id"] == "4481118"
    assert orders[0]["lines"][0]["item_name"] == "Oaty Breakfast Cookie"
