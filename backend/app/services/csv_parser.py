from __future__ import annotations

from io import BytesIO, StringIO
import csv
import io
from typing import Any, Dict, List, Optional, Tuple

from .unify_service import (
    CSV_ORDER_ID_FIELDS,
    UnifyService,
    UnifyServiceError,
)

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

try:  # pragma: no cover - optional dependency
    import xlrd
except Exception:  # pragma: no cover - optional dependency
    xlrd = None


_CSV_PARSER = UnifyService(base_url="", client_id="", client_secret="")


def _normalize_header(value: Any) -> str:
    text = _CSV_PARSER._normalize_csv_header(value)
    return text


def _header_matches_order_id(row: List[Any]) -> bool:
    normalized_values = {
        _normalize_header(cell)
        for cell in row
        if cell is not None and str(cell).strip()
    }
    required_values = {_normalize_header(field) for field in CSV_ORDER_ID_FIELDS}
    return any(value in required_values for value in normalized_values)


def _find_header_row(rows: List[List[Any]]) -> Optional[int]:
    for index, row in enumerate(rows[:20]):
        if _header_matches_order_id(row):
            return index
    return None


def _sheet_name_matches_orders_detailed(sheet_name: str) -> bool:
    normalized = _normalize_header(sheet_name)
    return "orders_detailed" in normalized or normalized == "ordersdetailed"


def _rows_to_csv_text(rows: List[List[Any]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    for row in rows:
        writer.writerow(["" if cell is None else cell for cell in row])
    return buffer.getvalue()


def _reader_to_rows(reader: csv.reader) -> List[List[Any]]:
    return [list(row) for row in reader]


def _load_csv_rows(content: bytes) -> Tuple[List[List[Any]], csv.Dialect]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    rows = _reader_to_rows(csv.reader(io.StringIO(text), dialect=dialect))
    return rows, dialect


def _load_xlsx_rows(content: bytes, filename: str) -> List[List[Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to parse Excel files")

    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    candidate_rows: List[List[Any]] = []

    preferred_sheet_name = next((sheet_name for sheet_name in workbook.sheetnames if _sheet_name_matches_orders_detailed(sheet_name)), None)
    sheet_names_to_scan = [preferred_sheet_name] if preferred_sheet_name else []
    sheet_names_to_scan.extend([sheet_name for sheet_name in workbook.sheetnames if sheet_name not in sheet_names_to_scan])

    for sheet_name in sheet_names_to_scan:
        worksheet = workbook[sheet_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if _find_header_row(rows) is not None:
            candidate_rows = rows
            break

    if not candidate_rows:
        candidate_sheet = workbook[workbook.sheetnames[0]]
        candidate_rows = [list(row) for row in candidate_sheet.iter_rows(values_only=True)]

    if not candidate_rows:
        raise UnifyServiceError(f"Excel file {filename} is empty")

    header_index = _find_header_row(candidate_rows)
    if header_index is None:
        raise UnifyServiceError("Could not find a CSV/XLSX header row with order_id columns")

    return candidate_rows[header_index:]


def _load_xls_rows(content: bytes, filename: str) -> List[List[Any]]:
    if xlrd is None:
        raise RuntimeError("xlrd is required to parse .xls files")

    workbook = xlrd.open_workbook(file_contents=content)
    if workbook.nsheets <= 0:
        raise UnifyServiceError(f"Excel file {filename} is empty")

    candidate_rows: List[List[Any]] = []
    sheets = workbook.sheets()
    preferred_sheets = [sheet for sheet in sheets if _sheet_name_matches_orders_detailed(sheet.name)]
    scan_sheets = preferred_sheets + [sheet for sheet in sheets if sheet not in preferred_sheets]

    for sheet in scan_sheets:
        rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
        if _find_header_row(rows) is not None:
            candidate_rows = rows
            break

    if not candidate_rows:
        sheet = workbook.sheet_by_index(0)
        candidate_rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]

    if not candidate_rows:
        raise UnifyServiceError(f"Excel file {filename} is empty")

    header_index = _find_header_row(candidate_rows)
    if header_index is None:
        raise UnifyServiceError("Could not find a CSV/XLSX header row with order_id columns")

    return candidate_rows[header_index:]


def _parse_csv_text(csv_text: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    orders, debug = _CSV_PARSER.parse_unify_csv_orders(csv_text)
    return [order.model_dump() for order in orders], debug


def parse_unify_file_bytes(content: bytes, filename: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Parse CSV, XLSX or XLS bytes and return JSON-safe orders plus debug summary."""
    lower = filename.lower()

    if lower.endswith(".csv") or lower.endswith(".txt") or lower.endswith(".tsv"):
        rows, _dialect = _load_csv_rows(content)
        header_index = _find_header_row(rows)
        if header_index is None:
            raise UnifyServiceError("CSV is missing an order_id column")
        csv_text = _rows_to_csv_text(rows[header_index:])
        return _parse_csv_text(csv_text)

    if lower.endswith(".xlsx"):
        rows = _load_xlsx_rows(content, filename)
        csv_text = _rows_to_csv_text(rows)
        return _parse_csv_text(csv_text)

    if lower.endswith(".xls"):
        rows = _load_xls_rows(content, filename)
        csv_text = _rows_to_csv_text(rows)
        return _parse_csv_text(csv_text)

    raise UnifyServiceError("Unsupported file type")
